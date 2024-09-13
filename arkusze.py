from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


workbook = Workbook()
sheet = workbook.active

# sheet["A1"] = "imię"
# sheet["B1"] = "Nazwisko"
#
# sheet["A2"] = "Jan"
# sheet["B2"] = "Kowalski"

# sheet.append(["Imię","nAZWISKO","miasto"])
# sheet.append(["Anna","Nowak","Ełk"])
#
# workbook.save("arkusz1.xlsx")

cities_data = [
    {
        "miasto": "Warszawa",
        "kraj": "Polska",
        "województwo": "Mazowieckie",
        "populacja": 1793579
    },
    {
        "miasto": "Kraków",
        "kraj": "Polska",
        "województwo": "Małopolskie",
        "populacja": 779115
    },
    {
        "miasto": "Wrocław",
        "kraj": "Polska",
        "województwo": "Dolnośląskie",
        "populacja": 643782
    },
    {
        "miasto": "Gdańsk",
        "kraj": "Polska",
        "województwo": "Pomorskie",
        "populacja": 486022
    },
    {
        "miasto": "Poznań",
        "kraj": "Polska",
        "województwo": "Wielkopolskie",
        "populacja": 534813
    },
    {
        "miasto": "Łódź",
        "kraj": "Polska",
        "województwo": "Łódzkie",
        "populacja": 672185
    },
    {
        "miasto": "Szczecin",
        "kraj": "Polska",
        "województwo": "Zachodniopomorskie",
        "populacja": 398255
    },
    {
        "miasto": "Rzeszów",
        "kraj": "Polska",
        "województwo": "Podkarpackie",
        "populacja": 196208
    }
]

def create_excel ():
    h = list(cities_data[0].keys())
    headers = list(map(lambda x: x.capitalize(), h))
    sheet.append(headers)
    for city in cities_data:
        row = list(city.values())
        sheet.append(row)

    gray_fill = PatternFill(start_color="DDDDDD", end_color="FFFFFF", fill_type="solid")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        if row[0].row % 2 == 0:
            for ceil in row:
                ceil.fill = gray_fill

    workbook.save("miasta.xlsx")


def read_excel():
    file = load_workbook("miasta.xlsx")
    sheet = file.active
    for row in sheet.iter_rows(values_only=True):
        print(sheet.iter_rows())

read_excel()







